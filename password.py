#coding:utf8
import os
import openpyxl

def file_check(file_check01):#check file information and build file
    if file_check01:
        os.chdir(password_pass + '\\secret')
    else:
        os.mkdir(password_pass + '\\secret')
        file_check02 = os.path.exists('secret')
        file_check(file_check02)

def person():#build personal document

password_pass = os.getcwd()
file_check01 = os.path.exists('secret')
file_check(file_check01)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'test_sheet1'
ws['A1'] = 'website'
ws['B1'] = 'person'
ws['C1'] = 'password'
data =
