import openpyxl
import os

a = os.path.exists("Empty_book.xlsx")
if  a:
    wb = openpyxl.load_workbook('Empty_book.xlsx')
    dest_filename = "Empty_book.xlsx"
else:
    wb = openpyxl.Workbook()
    dest_filename = "Empty_book.xlsx"
ws = wb.active
ws.title = 'range names'
ws['A1'] = 42
i = 0
while i<3:

    i++
ws.append([1,2,3,4])
wb.save(dest_filename)